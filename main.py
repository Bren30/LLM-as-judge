"""
main.py - Entry point for the Insight Quality Judge POC.

Workflow per run:
    1. Read input rows from the "Input" sheet in evaluations.xlsx.
    2. For each row:
       a. Load only the documents listed in doc_a_usar.
       b. Send the question, source text, and both answers to the judge agent.
       c. Parse the agent's JSON verdict.
       d. Append one row to the "Evaluations" sheet (with token counts).

Run from the project root:
    python main.py
"""

from __future__ import annotations

import asyncio
import io
import json
import os
import sys
from datetime import datetime
from pathlib import Path

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")

_PROJECT = os.environ.get("GOOGLE_CLOUD_PROJECT", "")
if _PROJECT != "acpe-dev-uc-ai":
    sys.exit(
        f"[SAFETY] Wrong GCP project detected: '{_PROJECT}'. "
        "This POC must only run against 'acpe-dev-uc-ai'. "
        "Check your .env file."
    )

from rich.console import Console
from rich.panel import Panel
from rich.rule import Rule
from rich.table import Table

from google.adk.runners import Runner
from google.adk.sessions import InMemorySessionService
from google.genai import types as genai_types

from prompt_judge.agent import root_agent

console = Console()

PROJECT_ROOT = Path(__file__).parent
DOCUMENTS_DIR = PROJECT_ROOT / "documents"
EXCEL_PATH = PROJECT_ROOT / "evaluations.xlsx"

EVAL_SHEET = "Evaluations"
INPUT_SHEET = "Input"

EXCEL_COLUMNS = [
    "fecha",
    "hora",
    "pregunta",
    "archivos_usados",
    "ganador",
    "margen",
    "puntaje_betty",
    "puntaje_copilot",
    "puntajes_por_dimension_betty",
    "puntajes_por_dimension_copilot",
    "dimension_mas_debil_betty",
    "dimension_mas_debil_copilot",
    "razonamiento",
    "mejora_principal",
    "respuesta_betty",
    "respuesta_copilot",
    "token_input",
    "token_output",
]

INPUT_COLUMNS = [
    "pregunta",
    "respuesta_betty",
    "respuesta_copilot",
    "doc_a_usar",
]

DIMENSION_SHORT_LABELS = {
    "Fidelidad y Fundamentacion": "Fidelidad",
    "Atribucion de Fuentes": "Atribucion",
    "Profundidad Analitica": "Profundidad",
    "Especificidad y Evidencia": "Especificidad",
    "Completitud": "Completitud",
    "Accionabilidad": "Accionabilidad",
    "Claridad y Transparencia del Razonamiento": "Claridad",
}


def format_per_dim(scores: dict | None) -> str:
    if not isinstance(scores, dict):
        return ""
    parts = []
    for full_name, short_name in DIMENSION_SHORT_LABELS.items():
        if full_name in scores:
            parts.append(f"{short_name}: {scores[full_name]}")
    return " | ".join(parts)


_SESSION_COUNTER = 0


# ---------------------------------------------------------------------------
# Document loading (PDF + DOCX only)
# ---------------------------------------------------------------------------

def _read_pdf(path: Path) -> str:
    from pypdf import PdfReader
    reader = PdfReader(str(path))
    parts = []
    for i, page in enumerate(reader.pages, 1):
        text = page.extract_text() or ""
        parts.append(f"[Page {i}]\n{text}")
    return "\n\n".join(parts)


def _read_docx(path: Path) -> str:
    from docx import Document
    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def load_source_documents(filenames: list[str] | None = None) -> tuple[str, list[str]]:
    """
    Load supported files in ./documents/ and concatenate the text.

    Args:
        filenames: If provided, only load files whose names are in this list.
                    If None, load all supported files.

    Returns:
        (combined_text, list_of_filenames)
    """
    if not DOCUMENTS_DIR.exists():
        sys.exit(
            f"[ERROR] Source folder not found: {DOCUMENTS_DIR}\n"
            "Create the folder and place your PDF/DOCX files inside before running."
        )

    all_files = sorted(
        p for p in DOCUMENTS_DIR.iterdir()
        if p.is_file() and p.suffix.lower() in (".pdf", ".docx")
    )

    if filenames:
        name_map = {f.name: f for f in all_files}
        selected = []
        for name in filenames:
            name_stripped = name.strip()
            if name_stripped not in name_map:
                sys.exit(
                    f"[ERROR] Document '{name_stripped}' not found in {DOCUMENTS_DIR}.\n"
                    f"Available: {', '.join(name_map.keys())}"
                )
            selected.append(name_map[name_stripped])
    else:
        selected = all_files

    if not selected:
        sys.exit(
            f"[ERROR] No PDF or DOCX files found matching the selection.\n"
            f"Available: {', '.join(f.name for f in all_files)}"
        )

    sections: list[str] = []
    used_names: list[str] = []
    for path in selected:
        try:
            if path.suffix.lower() == ".pdf":
                text = _read_pdf(path)
            else:
                text = _read_docx(path)
        except Exception as e:
            sys.exit(f"[ERROR] Could not read {path.name}: {e}")
        sections.append(f"===== FILE: {path.name} =====\n{text}")
        used_names.append(path.name)

    return "\n\n".join(sections), used_names


# ---------------------------------------------------------------------------
# Agent runner
# ---------------------------------------------------------------------------

def build_judge_input(
    question: str,
    source_excerpt: str,
    answer_betty: str,
    answer_copilot: str,
) -> str:
    return (
        f"## Question\n{question}\n\n"
        f"## Source Excerpt (from the documents both AIs received)\n"
        f"```\n{source_excerpt}\n```\n\n"
        f"## Answer from Betty\n```\n{answer_betty}\n```\n\n"
        f"## Answer from Copilot\n```\n{answer_copilot}\n```\n\n"
        "Evaluate both answers using the six insight-quality dimensions and "
        "return the JSON verdict only."
    )


async def run_judge(
    question: str,
    source_excerpt: str,
    answer_betty: str,
    answer_copilot: str,
) -> tuple[str, int | None, int | None]:
    global _SESSION_COUNTER
    _SESSION_COUNTER += 1
    session_id = f"eval_session_{_SESSION_COUNTER:03d}"

    session_service = InMemorySessionService()
    await session_service.create_session(
        app_name="insight_quality_judge",
        user_id="poc_user",
        session_id=session_id,
    )

    runner = Runner(
        agent=root_agent,
        app_name="insight_quality_judge",
        session_service=session_service,
    )

    user_message = genai_types.Content(
        role="user",
        parts=[genai_types.Part(
            text=build_judge_input(question, source_excerpt, answer_betty, answer_copilot)
        )],
    )

    final_response = ""
    token_input: int | None = None
    token_output: int | None = None

    async for event in runner.run_async(
        user_id="poc_user",
        session_id=session_id,
        new_message=user_message,
    ):
        if event.is_final_response():
            final_response = event.content.parts[0].text

        if event.usage_metadata is not None:
            um = event.usage_metadata
            if um.prompt_token_count is not None:
                token_input = (token_input or 0) + um.prompt_token_count
            if um.candidates_token_count is not None:
                token_output = (token_output or 0) + um.candidates_token_count

    return final_response, token_input, token_output


# ---------------------------------------------------------------------------
# JSON parsing
# ---------------------------------------------------------------------------

def parse_verdict(raw: str) -> dict:
    text = raw.strip()
    if text.startswith("```"):
        text = text.strip("`")
        if text.lower().startswith("json"):
            text = text[4:]
        text = text.strip("`").strip()
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1:
        sys.exit(f"[ERROR] Agent did not return JSON. Raw output:\n{raw}")
    try:
        return json.loads(text[start:end + 1])
    except json.JSONDecodeError as e:
        sys.exit(f"[ERROR] Could not parse agent JSON: {e}\nRaw:\n{raw}")


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _get_or_create_workbook():
    from openpyxl import Workbook, load_workbook

    if EXCEL_PATH.exists():
        return load_workbook(EXCEL_PATH)
    wb = Workbook()
    wb.active.title = EVAL_SHEET
    ws = wb[EVAL_SHEET]
    ws.append(EXCEL_COLUMNS)
    ws = wb.create_sheet(INPUT_SHEET)
    ws.append(INPUT_COLUMNS)
    wb.save(EXCEL_PATH)
    return wb


def ensure_input_sheet():
    """Ensure the Input sheet exists with headers. Returns the workbook."""
    from openpyxl import Workbook, load_workbook

    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        if INPUT_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(INPUT_SHEET)
            ws.append(INPUT_COLUMNS)
            wb.save(EXCEL_PATH)
        if EVAL_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(EVAL_SHEET)
            ws.append(EXCEL_COLUMNS)
            wb.save(EXCEL_PATH)
        return wb
    else:
        wb = Workbook()
        wb.active.title = EVAL_SHEET
        wb[EVAL_SHEET].append(EXCEL_COLUMNS)
        ws = wb.create_sheet(INPUT_SHEET)
        ws.append(INPUT_COLUMNS)
        wb.save(EXCEL_PATH)
        return wb


def read_input_rows() -> list[dict]:
    """Read all data rows from the Input sheet."""
    from openpyxl import load_workbook

    if not EXCEL_PATH.exists():
        sys.exit(
            f"[ERROR] Excel file not found: {EXCEL_PATH}\n"
            f"Create it with an '{INPUT_SHEET}' sheet containing columns: "
            f"{', '.join(INPUT_COLUMNS)}"
        )

    wb = load_workbook(EXCEL_PATH, data_only=True)
    if INPUT_SHEET not in wb.sheetnames:
        sys.exit(
            f"[ERROR] Sheet '{INPUT_SHEET}' not found in {EXCEL_PATH}.\n"
            f"Available sheets: {', '.join(wb.sheetnames)}"
        )

    ws = wb[INPUT_SHEET]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        vals = [ws.cell(row=row_idx, column=c).value for c in range(1, len(INPUT_COLUMNS) + 1)]
        pregunta = vals[0]
        if pregunta is None or str(pregunta).strip() == "":
            continue
        row_data = {}
        for i, col_name in enumerate(INPUT_COLUMNS):
            row_data[col_name] = vals[i] if i < len(vals) else ""
        rows.append(row_data)

    return rows


def append_to_excel(row: dict) -> None:
    from openpyxl import load_workbook

    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = _get_or_create_workbook()

    if EVAL_SHEET in wb.sheetnames:
        ws = wb[EVAL_SHEET]
    else:
        ws = wb.create_sheet(EVAL_SHEET)
        ws.append(EXCEL_COLUMNS)

    ws.append([row.get(col, "") for col in EXCEL_COLUMNS])
    wb.save(EXCEL_PATH)


# ---------------------------------------------------------------------------
# Main flow
# ---------------------------------------------------------------------------

def print_banner() -> None:
    console.print(
        Panel.fit(
            "[bold cyan]*** Insight Quality Judge - Batch Mode ***[/bold cyan]\n"
            "[dim]Betty vs Copilot - LLM-as-Judge for AI Engineering[/dim]\n"
            f"[dim]Project: {_PROJECT}[/dim]",
            border_style="cyan",
        )
    )


async def _async_main() -> None:
    print_banner()

    ensure_input_sheet()
    input_rows = read_input_rows()

    if not input_rows:
        console.print("[bold red]No input rows found in the 'Input' sheet.[/bold red]")
        console.print(f"[dim]Add rows with columns: {', '.join(INPUT_COLUMNS)}[/dim]")
        return

    console.print(Rule(f"[bold]Found {len(input_rows)} input row(s) to process[/bold]"))

    input_table = Table(title="Input Rows", show_lines=True)
    input_table.add_column("#", style="cyan", justify="right")
    input_table.add_column("Pregunta", style="white", max_width=40)
    input_table.add_column("Doc(s)", style="green", max_width=30)
    for i, row in enumerate(input_rows, 1):
        pregunta = str(row.get("pregunta", ""))[:40]
        docs = str(row.get("doc_a_usar", ""))[:30]
        input_table.add_row(str(i), pregunta, docs)
    console.print(input_table)

    for idx, row_data in enumerate(input_rows, 1):
        pregunta = str(row_data.get("pregunta", "")).strip()
        answer_betty = str(row_data.get("respuesta_betty", "")).strip()
        answer_copilot = str(row_data.get("respuesta_copilot", "")).strip()
        doc_a_usar_raw = str(row_data.get("doc_a_usar", "")).strip()

        if not pregunta:
            console.print(f"[yellow]Row {idx}: Empty pregunta, skipping.[/yellow]")
            continue

        doc_filenames = [d.strip() for d in doc_a_usar_raw.split(",") if d.strip()] if doc_a_usar_raw else []

        console.print(Rule(f"[bold]Row {idx}/{len(input_rows)}[/bold]"))
        console.print(f"[bold]Question:[/bold] {pregunta}")
        console.print(f"[bold]Documents:[/bold] {', '.join(doc_filenames) if doc_filenames else 'all'}")

        source_text, filenames = load_source_documents(
            filenames=doc_filenames if doc_filenames else None
        )
        console.print(f"[green]Loaded {len(filenames)} file(s):[/green] {', '.join(filenames)}")

        console.print("[yellow]Evaluating... please wait ~30 s[/yellow]")
        raw, token_input, token_output = await run_judge(pregunta, source_text, answer_betty, answer_copilot)
        verdict = parse_verdict(raw)

        now = datetime.now()
        result_row = {
            "fecha": now.strftime("%Y-%m-%d"),
            "hora": now.strftime("%H:%M:%S"),
            "pregunta": pregunta,
            "archivos_usados": ", ".join(filenames),
            "ganador": verdict.get("winner", ""),
            "margen": verdict.get("margin", ""),
            "puntaje_betty": verdict.get("score_betty", ""),
            "puntaje_copilot": verdict.get("score_copilot", ""),
            "puntajes_por_dimension_betty": format_per_dim(verdict.get("scores_betty_per_dim")),
            "puntajes_por_dimension_copilot": format_per_dim(verdict.get("scores_copilot_per_dim")),
            "dimension_mas_debil_betty": verdict.get("betty_weakest_dim", ""),
            "dimension_mas_debil_copilot": verdict.get("copilot_weakest_dim", ""),
            "razonamiento": verdict.get("reasoning", ""),
            "mejora_principal": verdict.get("top_improvement", ""),
            "respuesta_betty": answer_betty,
            "respuesta_copilot": answer_copilot,
            "token_input": token_input if token_input is not None else "",
            "token_output": token_output if token_output is not None else "",
        }
        append_to_excel(result_row)

        winner = verdict.get("winner", "?")
        if winner == "Empate":
            console.print(
                f"[bold]Resultado:[/bold] Empate - "
                f"(Betty {verdict.get('score_betty')} / Copilot {verdict.get('score_copilot')})"
            )
        else:
            console.print(
                f"[bold]Ganador:[/bold] {winner} "
                f"(Betty {verdict.get('score_betty')} / Copilot {verdict.get('score_copilot')}, "
                f"margen {verdict.get('margin')})"
            )
        console.print(f"[dim]Tokens: input={token_input}, output={token_output}[/dim]")
        console.print(f"[dim]Row {idx} saved to {EXCEL_PATH}[/dim]")

    console.print(Rule("[bold green]All rows processed![/bold green]"))
    console.print(f"[bold green]{len(input_rows)} evaluation(s) saved to {EXCEL_PATH}[/bold green]")


if __name__ == "__main__":
    try:
        asyncio.run(_async_main())
    except KeyboardInterrupt:
        console.print("\n[dim]Interrupted.[/dim]")